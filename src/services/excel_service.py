from openpyxl import load_workbook
# import datetime # 原來的匯入
from datetime import datetime # 明確匯入 datetime 類別
# from app.utils import western_to_roc_year # 移除錯誤的匯入
import logging
import os # 新增匯入 os
from openpyxl.styles import Alignment, Font

# --- 設定 ---
TEMPLATE_PATH = 'data/VSduty_template.xlsx' # 更新模板路徑
OUTPUT_DIR = 'data/output/' # 更新輸出目錄

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - [%(funcName)s] - %(message)s')
logger = logging.getLogger(__name__)

# --- 直接定義所需的函數 ---
def western_to_roc_year(western_year: int) -> int:
    """將西元年份轉換為民國年份"""
    return western_year - 1911
# ------------------------

class ExcelService:
    def __init__(self, template_path=TEMPLATE_PATH, output_dir=OUTPUT_DIR):
        """初始化 ExcelService。

        Args:
            template_path (str): Excel 模板檔案的路徑。
            output_dir (str): 儲存產生的 Excel 檔案的目錄。
        """
        self.template_path = template_path
        self.output_dir = output_dir

        # 確保輸出目錄存在
        # 檢查相對路徑是否存在
        if not os.path.exists(self.output_dir):
            # 嘗試相對於此腳本的父目錄的父目錄 (即專案根目錄) 建立
            script_dir = os.path.dirname(__file__) 
            potential_output_dir = os.path.abspath(os.path.join(script_dir, '..', '..', self.output_dir)) 
            logger.debug(f"Output directory {self.output_dir} not found relative to current working dir. Trying {potential_output_dir}")
            self.output_dir = potential_output_dir # 無論是否存在，都使用這個計算出的絕對路徑
            try:
                os.makedirs(self.output_dir, exist_ok=True)
                logger.info(f"確保輸出目錄存在: {self.output_dir}")
            except OSError as e:
                logger.error(f"建立輸出目錄失敗: {self.output_dir}, 錯誤: {e}")
                raise # 如果無法建立目錄，則拋出錯誤
        else:
             logger.info(f"輸出目錄已存在: {self.output_dir}")

         # 檢查模板檔案路徑
        if not os.path.exists(self.template_path):
            script_dir = os.path.dirname(__file__)
            potential_template_path = os.path.abspath(os.path.join(script_dir, '..', '..', self.template_path))
            logger.debug(f"Template file {self.template_path} not found relative to current working dir. Trying {potential_template_path}")
            if os.path.exists(potential_template_path):
                 self.template_path = potential_template_path
            else:
                 logger.error(f"錯誤：找不到模板檔案 {self.template_path} 或 {potential_template_path}")
                 raise FileNotFoundError(f"找不到模板檔案: {self.template_path}")

    def generate_excel(self, member_info: dict, duties: list[dict], year_month: str) -> tuple[str, str]:
        """根據成員資訊和值班記錄產生 Excel 檔案。

        Args:
            member_info (dict): 包含成員 'name' 和 'employee_id' 的字典。
            duties (list[dict]): 包含值班記錄的列表，每個記錄是一個字典，
                                包含 'date', 'weekday', 'start', 'end', 'work_hours', 'reason'。
            year_month (str): 年月字串 (YYYYMM)。

        Returns:
            tuple[str, str]: 包含產生的檔案路徑和用於下載的相對 URL 的元組。
                             如果生成失敗則返回 (None, None)。
        """
        employee_name = member_info.get('name', '未知')
        employee_id = member_info.get('employee_id', '未知')
        logger.info(f"開始為 {employee_name} ({employee_id}) 產生 {year_month} 的 Excel 報表...")

        try:
            # 載入模板
            workbook = load_workbook(self.template_path)
            sheet = workbook.active

            # --- 新增：處理合併儲存格 ---
            # 記錄所有合併儲存格範圍
            merged_ranges = list(sheet.merged_cells.ranges)
            # 暫時解除合併
            for merged_range in merged_ranges:
                logger.debug(f"Unmerging range: {merged_range}")
                sheet.unmerge_cells(str(merged_range))
            # --------------------------

            # 填寫標頭資訊
            sheet['C3'] = employee_name
            sheet['C4'] = employee_id
            sheet['G3'] = f"{year_month[:4]}年 {year_month[4:]}月"

            # 填寫值班資料 (從第 8 列開始)
            start_row = 8
            row_num = start_row
            total_work_hours = [0.0] * 5 # E, F, G, H, I 欄總計

            for duty in duties:
                if row_num > 38: # 假設模板最多填到第 38 列
                    logger.warning(f"資料筆數 ({len(duties)}) 超過模板容量 (31 筆)，僅填寫前 31 筆。")
                    break
                
                # 格式化日期
                try:
                     date_obj = datetime.strptime(duty['date'], "%Y%m%d")
                     formatted_date = f"{date_obj.month}/{date_obj.day}"
                except ValueError:
                     formatted_date = duty['date'] # 如果格式不對，保留原樣

                sheet[f'A{row_num}'] = formatted_date
                sheet[f'B{row_num}'] = duty['weekday']
                sheet[f'C{row_num}'] = duty['start']
                sheet[f'D{row_num}'] = duty['end']
                
                # 填寫工時分類 E-I
                work_hours = duty.get('work_hours', [0.0]*5)
                if len(work_hours) == 5:
                    sheet[f'E{row_num}'] = work_hours[0] if work_hours[0] else ''
                    sheet[f'F{row_num}'] = work_hours[1] if work_hours[1] else ''
                    sheet[f'G{row_num}'] = work_hours[2] if work_hours[2] else ''
                    sheet[f'H{row_num}'] = work_hours[3] if work_hours[3] else ''
                    sheet[f'I{row_num}'] = work_hours[4] if work_hours[4] else ''
                    # 累計總工時
                    for i in range(5):
                         total_work_hours[i] += work_hours[i]
                else:
                     logger.warning(f"工時資料格式不正確 (應為 5 個元素的列表): {work_hours}，日期: {duty['date']}")
                     for col in ['E', 'F', 'G', 'H', 'I']:
                         sheet[f'{col}{row_num}'] = ''

                sheet[f'J{row_num}'] = duty['reason']
                row_num += 1
            
            # 填寫總計 (在第 39 列)
            total_row = 39
            sheet[f'E{total_row}'] = total_work_hours[0] if total_work_hours[0] else ''
            sheet[f'F{total_row}'] = total_work_hours[1] if total_work_hours[1] else ''
            sheet[f'G{total_row}'] = total_work_hours[2] if total_work_hours[2] else ''
            sheet[f'H{total_row}'] = total_work_hours[3] if total_work_hours[3] else ''
            sheet[f'I{total_row}'] = total_work_hours[4] if total_work_hours[4] else ''
            logger.info(f"Set totals: {total_work_hours}")

            # --- 新增：重新合併儲存格 ---
            for merged_range in merged_ranges:
                 logger.debug(f"Re-merging range: {merged_range}")
                 sheet.merge_cells(str(merged_range))
            # --------------------------

            # 儲存檔案
            filename = f"{year_month}_{employee_name}.xlsx"
            output_path = os.path.join(self.output_dir, filename)
            workbook.save(output_path)
            logger.info(f"Excel file generated: {output_path}")
            
            # 生成相對 URL (假設 output 目錄可以被 Web 伺服器訪問)
            relative_url = f"/download/{filename}" # 這裡需要 API 端配合處理下載

            return output_path, relative_url

        except FileNotFoundError:
             logger.error(f"錯誤：找不到模板檔案 {self.template_path}。")
             return None, None
        except Exception as e:
            logger.error(f"產生 Excel 檔案時發生錯誤: {e}", exc_info=True)
            return None, None

# --- 測試代碼 (可選) ---
if __name__ == '__main__':
    print("Testing ExcelService...")
    # 確保模板在 data/ 且 data/output 存在
    try:
        service = ExcelService()
        
        # 模擬資料
        test_member = {'name': '測試員', 'employee_id': 'T001'}
        test_duties = [
            {'date': '20250401', 'weekday': '二', 'start': '1600', 'end': '2400', 'work_hours': [2.0, 2.0, 4.0, 0.0, 0.0], 'reason': '10'},
            {'date': '20250402', 'weekday': '三', 'start': '0000', 'end': '0800', 'work_hours': [2.0, 2.0, 4.0, 0.0, 0.0], 'reason': '10'},
            {'date': '20250404', 'weekday': '五', 'start': '0800', 'end': '2400', 'work_hours': [0.0, 0.0, 0.0, 8.0, 8.0], 'reason': '10'},
            {'date': '20250405', 'weekday': '六', 'start': '0000', 'end': '0800', 'work_hours': [0.0, 0.0, 0.0, 8.0, 0.0], 'reason': '10'},
             {'date': '20250408', 'weekday': '二', 'start': '0730', 'end': '0800', 'work_hours': [0.5, 0.0, 0.0, 0.0, 0.0], 'reason': '2. 科會'},
        ]
        year_month_test = '202504'
        
        file_path, url = service.generate_excel(test_member, test_duties, year_month_test)
        
        if file_path:
            print(f"Successfully generated test file: {file_path}")
            print(f"Relative URL: {url}")
        else:
            print("Failed to generate test file.")
            
    except Exception as e:
        print(f"An error occurred during testing: {e}")
        import traceback
        traceback.print_exc()
        
    print("Test finished.")


