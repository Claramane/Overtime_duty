from openpyxl import load_workbook
# import datetime # 原來的匯入
import datetime # 改回舊版匯入
# from app.utils import western_to_roc_year # 移除錯誤的匯入
import logging
import os # 新增匯入 os
from openpyxl.styles import Alignment, Font, Border, Side

# --- 設定 ---
# TEMPLATE_PATH = 'data/templates/VSduty_template.xlsx' # 改回舊版模板路徑
# TEMPLATE_PATH = 'data/VSduty_template.xlsx' # 使用者確認此路徑正確
TEMPLATE_PATH = 'VSduty_template.xlsx' # 僅使用文件名，將在初始化時計算完整路徑
# OUTPUT_DIR = 'data/output/' # 更新輸出目錄 - 保留新版的相對路徑處理和目錄建立
OUTPUT_DIR = 'output' # 僅使用目錄名，將在初始化時計算完整路徑

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - [%(funcName)s] - %(message)s')
logger = logging.getLogger(__name__)

# --- 直接定義所需的函數 ---
def western_to_roc_year(western_year: int) -> int:
    """將西元年份轉換為民國年份"""
    return western_year - 1911
# ------------------------

class ExcelService:
    def __init__(self, template_path=TEMPLATE_PATH, output_dir=OUTPUT_DIR):
        """初始化 ExcelService。

        Args:
            template_path (str): Excel 模板檔案的路徑。
            output_dir (str): 儲存產生的 Excel 檔案的目錄。
        """
        self.template_path = template_path
        self.output_dir = output_dir

        # 處理模板文件路徑
        if os.path.isabs(self.template_path) and os.path.exists(self.template_path):
            # 如果是絕對路徑且存在，直接使用
            logger.debug(f"使用絕對路徑模板文件: {self.template_path}")
        else:
            # 嘗試在不同位置查找模板文件
            if not os.path.exists(self.template_path):
                # 嘗試在 data 子目錄查找
                data_dir_path = os.path.join(os.getcwd(), 'backend', 'data')
                potential_path = os.path.join(data_dir_path, os.path.basename(self.template_path))
                logger.debug(f"嘗試從 data 目錄載入模板: {potential_path}")
                
                if os.path.exists(potential_path):
                    self.template_path = potential_path
                else:
                    # 嘗試相對於服務檔案的位置載入
                    script_dir = os.path.dirname(__file__)
                    base_dir = os.path.abspath(os.path.join(script_dir, '..', '..'))
                    data_dir = os.path.join(base_dir, 'data')
                    potential_path = os.path.join(data_dir, os.path.basename(self.template_path))
                    logger.debug(f"嘗試從腳本相對路徑載入模板: {potential_path}")
                    
                    if os.path.exists(potential_path):
                        self.template_path = potential_path
                    else:
                        logger.error(f"錯誤：找不到模板檔案 {self.template_path}")
                        raise FileNotFoundError(f"找不到模板檔案: {self.template_path}")
        
        logger.info(f"最終使用的模板文件路徑: {self.template_path}")

        # 處理輸出目錄路徑
        if os.path.isabs(self.output_dir) and os.path.exists(self.output_dir):
            # 如果是絕對路徑且存在，直接使用
            logger.debug(f"使用絕對路徑輸出目錄: {self.output_dir}")
        else:
            # 嘗試在不同位置創建或查找輸出目錄
            if not os.path.exists(self.output_dir):
                # 嘗試在 data 子目錄查找
                data_dir_path = os.path.join(os.getcwd(), 'backend', 'data')
                potential_path = os.path.join(data_dir_path, self.output_dir)
                logger.debug(f"嘗試使用 data 子目錄下的輸出目錄: {potential_path}")
                
                # 嘗試創建輸出目錄
                try:
                    os.makedirs(potential_path, exist_ok=True)
                    self.output_dir = potential_path
                except OSError:
                    # 如果無法創建，嘗試使用相對於服務檔案的位置
                    script_dir = os.path.dirname(__file__)
                    base_dir = os.path.abspath(os.path.join(script_dir, '..', '..'))
                    data_dir = os.path.join(base_dir, 'data')
                    potential_path = os.path.join(data_dir, self.output_dir)
                    logger.debug(f"嘗試使用腳本相對路徑的輸出目錄: {potential_path}")
                    
                    try:
                        os.makedirs(potential_path, exist_ok=True)
                        self.output_dir = potential_path
                    except OSError as e:
                        logger.error(f"建立輸出目錄失敗: {potential_path}, 錯誤: {e}")
                        raise
            
        logger.info(f"最終使用的輸出目錄路徑: {self.output_dir}")

    def generate_excel(self, member_info: dict, duties: list[dict], year_month: str) -> tuple[str, str]:
        """根據成員資訊和值班記錄產生 Excel 檔案 (使用舊版邏輯)。

        Args:
            member_info (dict): 包含成員 'name' 和 'employee_id' 的字典。
            duties (list[dict]): 包含值班記錄的列表，每個記錄是一個字典，
                                包含 'date', 'weekday', 'start', 'end', 'work_hours', 'reason'。
            year_month (str): 年月字串 (YYYYMM)。

        Returns:
            tuple[str, str]: 包含產生的檔案路徑和用於下載的相對 URL 的元組。
                             如果生成失敗則返回 (None, None)。
        """
        employee_name = member_info.get('name', '未知')
        # employee_id = member_info.get('employee_id', '未知') # 舊版在 _set_member_info 中處理
        logger.info(f"開始為 {employee_name} 產生 {year_month} 的 Excel 報表 (舊版邏輯)...")
        filename = f"{year_month}_{employee_name}.xlsx" # 與舊版一致
        output_path = os.path.join(self.output_dir, filename)

        try:
            # 載入模板
            workbook = load_workbook(self.template_path)
            sheet = workbook.active

            # --- 處理合併儲存格 (保留新版邏輯) ---
            merged_ranges = list(sheet.merged_cells.ranges)
            for merged_range in merged_ranges:
                logger.debug(f"Unmerging range: {merged_range}")
                sheet.unmerge_cells(str(merged_range))
            # --------------------------

            # --- 使用舊版輔助方法填寫 ---
            western_year = int(year_month[:4])
            roc_year = western_to_roc_year(western_year)
            month = year_month[4:]
            header_text = f"{roc_year}年{month}月份主治醫師加班時數彙整表" # 舊版標頭

            self._set_header(sheet, header_text)
            self._set_member_info(sheet, member_info)

            # 舊版資料從第 5 列開始
            start_row = 5
            row_index = start_row
            # 舊版總計欄位包含總工時，共 6 個元素 (總時數 + 5 個分類時數)
            totals = [0.0] * 6

            for duty in duties:
                 # 舊版模板容量可能不同，這裡暫不限制行數，依賴模板設計
                 self._fill_duty_row(sheet, row_index, duty, totals)
                 row_index += 1

            # 舊版總計在第 20 列
            self._set_totals(sheet, totals)
            # --------------------------


            # --- 重新合併儲存格 (保留新版邏輯) ---
            for merged_range in merged_ranges:
                 logger.debug(f"Re-merging range: {merged_range}")
                 sheet.merge_cells(str(merged_range))
            # --------------------------

            # --- 對 A2:L2 加上粗框線 ---
            self._set_thick_border(sheet, 'A2:L2')
            # --------------------------

            # --- 對 K1:L1 (姓名欄位) 加上框線 ---
            thin_side = Side(style='thin')
            # K1: 上、下、左邊框
            sheet['K1'].border = Border(top=thin_side, bottom=thin_side, left=thin_side)
            # L1: 上、下、右邊框
            sheet['L1'].border = Border(top=thin_side, bottom=thin_side, right=thin_side)
            logger.info("已對 K1:L1 設定框線")
            # --------------------------

            # --- 取消 A1、D1、G1、J1 的粗體字 ---
            for cell_addr in ['A1', 'D1', 'G1', 'J1']:
                cell = sheet[cell_addr]
                cell.font = Font(bold=False)
            logger.info("已取消 A1、D1、G1、J1 的粗體字")
            # --------------------------

            # 儲存檔案
            workbook.save(output_path)
            logger.info(f"Excel file generated: {output_path}")

            # 生成相對 URL (與舊版一致)
            relative_url = f"/download/{filename}"

            return output_path, relative_url

        except FileNotFoundError:
             logger.error(f"錯誤：找不到模板檔案 {self.template_path}。")
             return None, None
        except Exception as e:
            logger.error(f"產生 Excel 檔案時發生錯誤: {e}", exc_info=True)
            return None, None

    # --- 舊版輔助方法 ---
    def _set_header(self, ws, header_text):
        """設定頁首 (置中顯示標題)

        Args:
            ws: 工作表物件
            header_text: 頁首文字，例如「115年01月份主治醫師加班時數彙整表」
        """
        # 設定頁首置中文字
        ws.oddHeader.center.text = header_text
        ws.oddHeader.left.text = ""
        ws.oddHeader.right.text = ""

        # 頁面設定
        ws.page_setup.differentFirst = False

        logger.info(f"已設定頁首文字: {header_text}")

    def _set_thick_border(self, ws, cell_range):
        """對指定範圍設定框線（適用於合併儲存格）

        Args:
            ws: 工作表物件
            cell_range: 儲存格範圍，例如 'A2:L2'
        """
        border_side = Side(style='thin')

        # 解析範圍
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)

                # 合併儲存格：每個儲存格都要設定上下邊框，左右只設定邊界
                top = border_side
                bottom = border_side
                left = border_side if col == min_col else None
                right = border_side if col == max_col else None

                cell.border = Border(top=top, bottom=bottom, left=left, right=right)

        logger.info(f"已對 {cell_range} 設定框線")

    def _set_member_info(self, ws, member):
        """填寫成員資訊 (舊版邏輯)"""
        ws['B1'] = "麻醉科" # 假設固定值
        ws['E1'] = "6200"   # 假設固定值
        ws['H1'] = member.get('employee_id', '未知')
        ws['K1'] = member.get('name', '未知')
        logger.info(f"已填寫成員資訊: {member.get('name')}")

    def _fill_duty_row(self, ws, row, duty, totals):
        """填寫單行值班資料 (舊版邏輯)"""
        try:
            date_obj = datetime.datetime.strptime(duty['date'], "%Y%m%d")
            formatted_date = f"{date_obj.month:02d}/{date_obj.day:02d}" # 舊版格式 MM/DD
        except ValueError:
            formatted_date = duty['date'] # 格式錯誤則保留原樣

        ws.cell(row=row, column=1, value=formatted_date) # A 欄
        ws.cell(row=row, column=2, value=duty.get('weekday', '')) # B 欄
        ws.cell(row=row, column=3, value=f"{duty.get('start', '')}-{duty.get('end', '')}") # C 欄: 合併時間

        work_hours = duty.get('work_hours', [0.0]*5)
        total_hours = 0.0
        if isinstance(work_hours, list) and len(work_hours) == 5:
             try:
                 total_hours = sum(float(h) for h in work_hours if h) # 計算總工時
             except (ValueError, TypeError):
                 logger.warning(f"無法計算第 {row} 行的總工時: {work_hours}")
                 total_hours = 0.0 # Default to 0 if calculation fails
        else:
             logger.warning(f"第 {row} 行工時資料格式不正確 (應為 5 個元素的列表): {work_hours}")
             work_hours = [0.0] * 5 # 使用預設值避免後續錯誤

        ws.cell(row=row, column=4, value=total_hours).number_format = '0.0' # D 欄: 總工時 (填入 0.0)
        totals[0] += total_hours  # Accumulate total hours to totals[0]

        # E-I 欄: 分類工時
        for i, hours in enumerate(work_hours, start=5): # 從第 5 欄開始
            cell_value = hours if hours else '' # 空值或 0 顯示為空
            # 直接使用 hours 值，確保 0.0 被寫入
            # 處理 hours 可能為 None 或非數字的情況
            numeric_hours = 0.0
            try:
                if hours is not None and hours != '': # 檢查非空
                    numeric_hours = float(hours)
            except (ValueError, TypeError):
                logger.warning(f"無法轉換第 {row} 行第 {i} 欄的工時為數字: {hours}")

            ws.cell(row=row, column=i, value=numeric_hours).number_format = '0.0'
            try:
                 totals[i-4] += float(hours) if hours else 0.0 # 累加到 totals[1] 到 totals[5]
            except (ValueError, TypeError):
                 logger.warning(f"無法累加第 {row} 行第 {i} 欄的工時: {hours}")

        ws.cell(row=row, column=10, value=duty.get('reason', '')) # J 欄: 事由

        logger.debug(f"已填寫第 {row} 行: 日期={formatted_date}, 時間={ws.cell(row=row, column=3).value}, 總時數={total_hours}, 事由={duty.get('reason', '')}")

    def _set_totals(self, ws, totals):
        """填寫總計列 (舊版邏輯，第 20 列)"""
        total_row = 20 # 舊版總計列
        # 舊版總計欄位是 D 到 I (column 4 到 9)
        for col in range(4, 10):
            # 直接寫入總計值，包括 0.0
            total_value_numeric = totals[col-4]
            cell = ws.cell(row=total_row, column=col, value=total_value_numeric)
            cell.number_format = '0.0' # 統一設定數字格式

        logger.info(f"已設定總計 (第 {total_row} 行): {[f'{t:.1f}' if isinstance(t, float) else t for t in totals]}") # Log formatted totals

# --- 測試代碼 (可選) ---
# if __name__ == '__main__':
#     print("Testing ExcelService (Old Logic Style)...")
#     # 確保模板在 data/templates/ 且 data/output 存在
#     try:
#         # 創建必要的目錄結構 (如果不在的話)
#         # if not os.path.exists('data/templates'):
#         #     os.makedirs('data/templates')
#         #     print("Created 'data/templates' directory for testing.")
#         #     # 注意：這裡需要您手動放入一個 VSduty_template.xlsx 檔案到 data/templates
#         #     print("Please place your 'VSduty_template.xlsx' in 'data/templates' for the test to run correctly.")
#         # # 為了測試，我們先假設模板存在，如果不存在，ExcelService 初始化會報錯
#         # if not os.path.exists('data/templates/VSduty_template.xlsx'):
#         #      print("WARNING: 'data/templates/VSduty_template.xlsx' not found. Creating a dummy one.")
#         #      from openpyxl import Workbook
#         #      dummy_wb = Workbook()
#         #      dummy_ws = dummy_wb.active
#         #      # 可以添加一些基本結構以防報錯，但最好是用真實模板
#         #      dummy_ws['A1'] = "範例標題"
#         #      # Add headers expected by _set_member_info to avoid errors if template is missing
#         #      dummy_ws['B1'] = "科別"
#         #      dummy_ws['E1'] = "代碼"
#         #      dummy_ws['H1'] = "員工ID"
#         #      dummy_ws['K1'] = "姓名"
#         #      # Add a few rows for data and the total row
#         #      for r in range(5, 21):
#         #          for c in range(1, 11):
#         #               dummy_ws.cell(row=r, column=c, value=f"R{r}C{c}")
#         #      dummy_ws.cell(row=20, column=3, value="總計") # Example total label
#         #      dummy_wb.save('data/templates/VSduty_template.xlsx')
#         # # 正確模板路徑應為 data/VSduty_template.xlsx
#         if not os.path.exists(TEMPLATE_PATH):
#              print(f"WARNING: Template {TEMPLATE_PATH} not found. Cannot run test effectively.")
#         else:
#             service = ExcelService() # 會使用設定的模板路徑
#             # 模擬資料 (與新版測試資料類似，但適配舊版邏輯)
#             test_member = {'name': '測試員舊版', 'employee_id': 'T001OLD'}
#             test_duties = [
#                 {'date': '20250401', 'weekday': '二', 'start': '1600', 'end': '2400', 'work_hours': [2.0, 2.0, 4.0, 0.0, 0.0], 'reason': '10'},
#                 {'date': '20250402', 'weekday': '三', 'start': '0000', 'end': '0800', 'work_hours': [2.0, 2.0, 4.0, 0.0, 0.0], 'reason': '10'},
#                 {'date': '20250404', 'weekday': '五', 'start': '0800', 'end': '2400', 'work_hours': [0.0, 0.0, 0.0, 8.0, 8.0], 'reason': '10'},
#                 {'date': '20250405', 'weekday': '六', 'start': '0000', 'end': '0800', 'work_hours': [0.0, 0.0, 0.0, 8.0, 0.0], 'reason': '10'},
#                  {'date': '20250408', 'weekday': '二', 'start': '0730', 'end': '0800', 'work_hours': [0.5, 0.0, 0.0, 0.0, 0.0], 'reason': '2. 科會'},
#                  # Add data with None/empty values to test handling
#                  {'date': '20250409', 'weekday': '三', 'start': '1000', 'end': '1200', 'work_hours': [None, 1.0, 1.0, None, ''], 'reason': '測試空值'},
#             ]
#             # 填充更多數據，測試超過20行總計的情況 (總計仍在第20行)
#             for i in range(10, 25):
#                  day = f"{i:02d}"
#                  # Ensure work_hours has 5 elements, even if some are 0
#                  test_duties.append({'date': f'202504{day}', 'weekday': 'X', 'start': '1200', 'end': '1300', 'work_hours': [1.0, 0.0, 0.0, 0.0, 0.0], 'reason': '午休加班'})
#             year_month_test = '202504'
#             file_path, url = service.generate_excel(test_member, test_duties, year_month_test)
#             if file_path:
#                 print(f"Successfully generated test file (old logic): {file_path}")
#                 print(f"Relative URL: {url}")
#             else:
#                 print("Failed to generate test file (old logic).")
#     except FileNotFoundError as e:
#          print(f"Error during testing: {e}")
#          print(f"Please ensure the template file exists at '{TEMPLATE_PATH}'")
#     except Exception as e:
#         print(f"An error occurred during testing: {e}")
#         import traceback
#         traceback.print_exc()
#     print("Test finished.")


